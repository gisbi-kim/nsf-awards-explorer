"""Extract all robotics-related awards across all FY slim files in data/slim/.
Output: data/nsf_robotics_awards_all_years.xlsx

Run after slim files are generated. Idempotent (re-runs scan all available years).
"""
import re
from pathlib import Path
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
SLIM_DIR = ROOT / "data" / "slim"
OUT = ROOT / "data" / "nsf_robotics_awards_all_years.xlsx"

# Strict robotics pattern — must contain robot* stem OR specific robotics terms.
# Excludes generic words ("navigation", "manipulation", "embodied", "drone/uav",
# "actuator", "grasping") that over-match in non-robotics fields (oceanography,
# chemistry, cognitive science, mechanical engineering).
ROBOT_PAT = re.compile(
    r"\b(?:"
    r"robot|robots|robotic|robotics|robotically|"
    r"slam|"
    r"human-robot|hri|"
    r"teleoperation|teleoperated|"
    r"legged locomotion|"
    r"autonomous (?:vehicle|vehicles|driving|mobility)|"
    r"self-driving|"
    r"unmanned aerial vehicle|"
    r"unmanned ground vehicle|"
    r"unmanned aircraft system|"
    r"motion planning|"
    r"robot learning|"
    r"manipulator arm|"
    r"robot manipulation"
    r")\b",
    re.IGNORECASE,
)

slim_files = sorted(SLIM_DIR.glob("nsf_awards_FY*_slim.xlsx"))
print(f"found slim files: {[f.name for f in slim_files]}")

dfs = []
for f in slim_files:
    fy = int(re.search(r"FY(\d{4})", f.name).group(1))
    df = pd.read_excel(f)
    df.insert(0, "fy", fy)
    dfs.append(df)
all_df = pd.concat(dfs, ignore_index=True)
print(f"\nTotal awards across all years: {len(all_df):,}")

all_df["abstract_lc"] = all_df["abstract"].fillna("").astype(str).str.lower()
all_df["title_lc"]    = all_df["title_clean"].fillna("").astype(str).str.lower()

mask_abs   = all_df["abstract_lc"].str.contains(ROBOT_PAT, regex=True, na=False)
mask_title = all_df["title_lc"].str.contains(ROBOT_PAT, regex=True, na=False)
mask = mask_abs | mask_title

print(f"  matched in title:    {mask_title.sum():,}")
print(f"  matched in abstract: {mask_abs.sum():,}")
print(f"  matched in either:   {mask.sum():,}")

rob = all_df[mask].copy()
keep = ["fy","award_id","award_url","program_prefix","title_clean",
        "pi_first","pi_last","pi_email","co_pi_names",
        "institution","state","directorate","directorate_full",
        "division","division_full","program_name",
        "po_name","po_email","amount_usd","transaction_type",
        "start_date","end_date","abstract"]
keep = [c for c in keep if c in rob.columns]

rob["matched_in"] = ""
rob.loc[mask_title & mask_abs, "matched_in"] = "title+abstract"
rob.loc[mask_title & ~mask_abs, "matched_in"] = "title only"
rob.loc[~mask_title & mask_abs, "matched_in"] = "abstract only"

rob = rob[keep + ["matched_in"]]
rob = rob.sort_values(["fy", "amount_usd"], ascending=[False, False])

LIMIT = 32_000
def _trunc(v):
    if isinstance(v, str) and len(v) > LIMIT:
        return v[:LIMIT] + " ...[T]"
    return v
rob = rob.apply(lambda c: c.map(_trunc))

yearly = rob.groupby("fy").agg(
    n_awards=("award_id","count"),
    total_M=("amount_usd", lambda s: pd.to_numeric(s, errors="coerce").sum()/1e6),
).round(1).reset_index().sort_values("fy", ascending=False)

inst_top = rob.groupby("institution").agg(
    n=("award_id","count"),
    total_M=("amount_usd", lambda s: pd.to_numeric(s, errors="coerce").sum()/1e6),
).sort_values("n", ascending=False).head(50).round(1).reset_index()

pi_top = rob.groupby(["pi_first","pi_last","pi_email","institution"]).agg(
    n=("award_id","count"),
    total_M=("amount_usd", lambda s: pd.to_numeric(s, errors="coerce").sum()/1e6),
).sort_values("n", ascending=False).head(50).round(1).reset_index()

OUT.parent.mkdir(parents=True, exist_ok=True)
with pd.ExcelWriter(OUT, engine="openpyxl") as w:
    rob.to_excel(w,      sheet_name="all_robotics_awards", index=False)
    yearly.to_excel(w,   sheet_name="yearly_summary",      index=False)
    inst_top.to_excel(w, sheet_name="top_institutions",    index=False)
    pi_top.to_excel(w,   sheet_name="top_pis",             index=False)

    from openpyxl.styles import Font, Alignment
    ws = w.sheets["all_robotics_awards"]
    if "award_url" in rob.columns:
        url_col = list(rob.columns).index("award_url") + 1
        link_font = Font(color="0563C1", underline="single")
        for r in range(2, len(rob) + 2):
            cell = ws.cell(row=r, column=url_col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = link_font
    if "abstract" in rob.columns:
        abs_col = list(rob.columns).index("abstract") + 1
        for r in range(2, len(rob) + 2):
            ws.cell(row=r, column=abs_col).alignment = Alignment(wrap_text=True, vertical="top")
    widths = {1:6,2:12,3:42,4:18,5:55,6:14,7:14,8:28,9:36,10:8,11:8,
              12:30,13:8,14:36,15:30,16:24,17:28,18:14,19:18,20:12,21:12,22:80,23:14}
    for col_idx, w_ in widths.items():
        if col_idx <= len(rob.columns):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = w_
    ws.freeze_panes = "B2"

print(f"\nwrote {OUT}")
print(f"  sheets: all_robotics_awards ({len(rob):,}), yearly_summary ({len(yearly)}), top_institutions ({len(inst_top)}), top_pis ({len(pi_top)})")
print(f"\nyearly summary:")
print(yearly.to_string(index=False))
